import os
import re
import tempfile
import uuid
from typing import Dict, List
from datetime import datetime
from datetime import timedelta
import openpyxl
from telegram import Update, InputFile, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, MessageHandler, CallbackQueryHandler, filters, ContextTypes

AVAILABLE_SOURCES = ["RANA", "SHA", "KAR", "BL", "CUSTOM"]

pending_sources: Dict[int, Dict] = {}
source_file_counter: Dict[str, int] = {}
current_day = (datetime.utcnow() + timedelta(hours=7)).strftime('%d%m')

class TelegramBot:
    def __init__(self, token: str):
        self.token = token
        self.application = Application.builder().token(token).build()
        
        self.application.add_handler(MessageHandler(filters.Document.ALL, self.handle_document))
        self.application.add_handler(CallbackQueryHandler(self.handle_source_selection, pattern=r'^source_'))

    async def start(self):
        await self.application.initialize()
        await self.application.start()
        await self.application.updater.start_polling()
    
    async def stop(self):
        await self.application.updater.stop()
        await self.application.stop()
        await self.application.shutdown()

    async def handle_document(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not update.message or not update.message.document:
            return
        
        message = update.message
        chat_id = message.chat_id
        doc = message.document
        
        if not (doc.file_name.lower().endswith('.txt') or doc.file_name.lower().endswith('.xlsx')):
            await message.reply_text("❌ Vui lòng gửi file TXT hoặc Excel (.xlsx).")
            return

        if doc.file_size > 10 * 1024 * 1024:
            await message.reply_text("❌ File quá lớn. Dung lượng tối đa là 10MB.")
            return
        
        try:
            file = await doc.get_file()
            temp_dir = os.path.join(tempfile.gettempdir(), "TelegramBotTemp")
            os.makedirs(temp_dir, exist_ok=True)
            temp_path = os.path.join(temp_dir, f"{uuid.uuid4()}_{doc.file_name}")
            
            await file.download_to_drive(temp_path)

            pending_sources[chat_id] = {
                "file_path": temp_path,
                "original_file_name": doc.file_name
            }

            keyboard = [
                [InlineKeyboardButton(f"{source}", callback_data=f"source_{source}") for source in AVAILABLE_SOURCES]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)

            await message.reply_text(
                "📥 Đã nhận file. Vui lòng chọn *source* bằng cách nhấn nút bên dưới:",
                reply_markup=reply_markup,
                parse_mode='Markdown'
            )
            
        except Exception as e:
            await message.reply_text(f"⚠️ Lỗi: {str(e)}")
            if 'temp_path' in locals() and os.path.exists(temp_path):
                os.remove(temp_path)

    async def handle_source_selection(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        await query.answer()
        chat_id = query.message.chat_id
        source_input = query.data.replace("source_", "").upper()

        global current_day
        today = (datetime.utcnow() + timedelta(hours=7)).strftime('%d%m')
        if today != current_day:
            source_file_counter.clear()
            current_day = today

        if chat_id not in pending_sources:
            await query.edit_message_text("❌ Không tìm thấy file chờ xử lý.")
            return

        pending_file = pending_sources.pop(chat_id)
        file_num = source_file_counter.get(source_input, 0) + 1
        source_file_counter[source_input] = file_num

        if pending_file["file_path"].lower().endswith('.xlsx'):
            preview_lines = await self.process_excel_file(pending_file["file_path"], "")
        else:
            preview_lines = await self.process_txt_file(pending_file["file_path"], "")
        
        gmail_count = len(preview_lines)
        time_str = (datetime.utcnow() + timedelta(hours=7)).strftime('%H%M%S')

        full_source = f"{source_input}_{today}_{file_num}_{gmail_count}-gmails_{time_str}"
        await query.edit_message_text(f"⏳ Đang xử lý với source: {full_source}")

        if pending_file["file_path"].lower().endswith('.xlsx'):
            processed_lines = await self.process_excel_file(pending_file["file_path"], full_source)
        else:
            processed_lines = await self.process_txt_file(pending_file["file_path"], full_source)

        if not processed_lines:
            await query.message.reply_text("❌ Không tìm thấy tài khoản Gmail nào trong file.")
            return

        processed_path = os.path.join(tempfile.gettempdir(), f"processed_{uuid.uuid4()}.txt")
        with open(processed_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(processed_lines))

        with open(processed_path, 'rb') as file_to_send:
            await query.message.reply_document(
                document=InputFile(file_to_send, filename=f"{full_source}.txt"),
                caption=f"✅ Đã xử lý {gmail_count} tài khoản Gmail\nSource: {full_source}"
            )

        os.remove(pending_file["file_path"])
        os.remove(processed_path)

    async def process_txt_file(self, path: str, source: str) -> List[str]:
        formatted = []
        with open(path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                parts = re.sub(r'\s+', ' ', line).split(' ')
                parts = [p for p in parts if p]
                if parts and "@gmail.com" in parts[0].lower():
                    if len(parts) == 1:
                        result = f"{parts[0]}|aass1122|SOURCE_{source}_SOURCE"
                    elif len(parts) == 2:
                        result = f"{parts[0]}|{parts[1]}|SOURCE_{source}_SOURCE"
                    else:
                        result = f"{parts[0]}|{parts[1]}|{parts[2]}|SOURCE_{source}_SOURCE"
                    formatted.append(result)
        return formatted

    async def process_excel_file(self, path: str, source: str) -> List[str]:
        lines = []
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, values_only=True):
            emails = []
            non_emails = []
            for cell in row[:5]:
                if cell:
                    text = str(cell).strip()
                    if '@' in text and '.' in text:
                        emails.append(text)
                    else:
                        non_emails.append(text)
            if emails:
                email = emails[0]
                recovery = emails[1] if len(emails) > 1 else ''
                password = non_emails[0] if non_emails else 'aass1122'
                line = f"{email}|{password}"
                if recovery:
                    line += f"|{recovery}"
                line += f"|SOURCE_{source}_SOURCE"
                lines.append(line)
        return lines
    # async def process_excel_file(self, path: str, source: str) -> List[str]:
    #     lines = []
    #     wb = openpyxl.load_workbook(path)
    #     ws = wb.active

    #     for row in ws.iter_rows(values_only=True):  # duyệt tất cả các dòng
    #         emails = []
    #         non_emails = []
    #         for cell in row[:5]:
    #             if cell:
    #                 text = str(cell).strip()
    #                 if '@' in text and '.' in text:
    #                     emails.append(text)
    #                 else:
    #                     non_emails.append(text)
    #         if emails:
    #             email = emails[0]
    #             recovery = emails[1] if len(emails) > 1 else ''
    #             password = non_emails[0] if non_emails else 'aass1122'
    #             line = f"{email}|{password}"
    #             if recovery:
    #                 line += f"|{recovery}"
    #             line += f"|SOURCE_{source}_SOURCE"
    #             lines.append(line)
    #     return lines


async def main():
    token = os.getenv("TELEGRAM_BOT_TOKEN")
    if not token:
        raise ValueError("Thiếu biến môi trường TELEGRAM_BOT_TOKEN")
    bot = TelegramBot(token)
    await bot.start()
    while True:
        await asyncio.sleep(1)

if __name__ == "__main__":
    import asyncio
    import dotenv
    dotenv.load_dotenv()
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("Bot đã dừng.")
